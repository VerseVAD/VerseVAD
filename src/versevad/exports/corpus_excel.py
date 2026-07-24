"""Readable, auditable Excel export for a completed local corpus project."""

from __future__ import annotations

from dataclasses import asdict
from io import BytesIO
import json
from typing import Iterable, Mapping, Sequence

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from versevad.corpus import (
    corpus_module_category_profiles,
    corpus_module_profiles,
    corpus_vad_profiles,
)
from versevad.db import (
    CorpusMetricRecord,
    CorpusModuleAggregateRecord,
    CorpusModuleCoverageRecord,
    CorpusModuleMetricRecord,
    CorpusModuleResultRecord,
    CorpusModuleWarningRecord,
    CorpusTextRecord,
    ProjectRecord,
    UnmatchedQcRecord,
)


INK = "172A3A"
RUST = "A34F32"
SAGE = "5F7661"
PAPER = "FBF8F1"
PALE = "EEF3EC"
WHITE = "FFFFFF"

CORPUS_WORKBOOK_API_VERSION = 5


def _label(name: str) -> str:
    return name.replace("_", " ").strip().title()


def _construct_label(metric: str, category: str) -> str:
    if metric == "association_rate":
        if category in {"positive", "negative"}:
            return "Sentiment Association"
        return "Emotion Association"
    if metric.startswith("intensity_"):
        return "Emotion Intensity"
    return "Coverage"


def _cell_value(value: object) -> object:
    if isinstance(value, (dict, list, tuple)):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _write_sheet(
    workbook: Workbook,
    *,
    title: str,
    purpose: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[object]],
    table_name: str,
) -> None:
    sheet = workbook.create_sheet(title)
    sheet.sheet_view.showGridLines = False
    sheet["A1"] = title
    sheet["A1"].font = Font(name="Aptos Display", size=18, bold=True, color=INK)
    sheet["A2"] = purpose
    sheet["A2"].font = Font(name="Aptos", size=10, italic=True, color=SAGE)
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 1))
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    header_row = 4
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(header_row, column, header)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=INK)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    materialized = tuple(rows)
    for row_number, values in enumerate(materialized, start=header_row + 1):
        for column, value in enumerate(values, start=1):
            cell = sheet.cell(row_number, column, value)
            cell.font = Font(name="Aptos", size=10, color=INK)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            header = headers[column - 1].casefold()
            if isinstance(value, float):
                cell.number_format = "0.000"
                if "coverage" in header or "rate" in header:
                    cell.number_format = "0.0%"
    last_row = header_row + len(materialized)
    if materialized:
        table = Table(
            displayName=table_name,
            ref=f"A{header_row}:{get_column_letter(len(headers))}{last_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        sheet.add_table(table)
    sheet.freeze_panes = "A5"
    sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(headers))}{max(last_row, header_row)}"
    for column, header in enumerate(headers, start=1):
        observed = [len(str(header))]
        observed.extend(len(str(row[column - 1] or "")) for row in materialized)
        width = min(max(max(observed) + 2, 11), 42)
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.row_dimensions[2].height = 34


def build_corpus_workbook(
    project: ProjectRecord,
    texts: Sequence[CorpusTextRecord],
    metrics: Sequence[CorpusMetricRecord],
    unmatched: Sequence[UnmatchedQcRecord],
    methodology: Mapping[str, object] | None = None,
    review_decisions: Sequence[Mapping[str, object]] = (),
    part_of_speech_rows: Sequence[Mapping[str, object]] = (),
    module_metrics: Sequence[CorpusModuleMetricRecord] = (),
    module_coverage: Sequence[CorpusModuleCoverageRecord] = (),
    module_results: Sequence[CorpusModuleResultRecord] = (),
    module_aggregates: Sequence[CorpusModuleAggregateRecord] = (),
    module_warnings: Sequence[CorpusModuleWarningRecord] = (),
) -> bytes:
    """Return an in-memory `.xlsx`; no project text is sent elsewhere."""

    workbook = Workbook()
    readme = workbook.active
    readme.title = "START HERE"
    readme.sheet_view.showGridLines = False
    readme["A1"] = "VerseVAD corpus workbook"
    readme["A1"].font = Font(name="Aptos Display", size=22, bold=True, color=INK)
    readme["A3"] = project.title
    readme["A3"].font = Font(name="Aptos Display", size=16, bold=True, color=RUST)
    guidance = (
        (
            "Begin with",
            "Corpus Profiles, then Work VAD, Language Profile, and Coverage and Emotion.",
        ),
        (
            "Part-of-speech profile",
            "Counts and shares use all eligible lexical tokens and are independent of affective-lexicon coverage.",
        ),
        (
            "Additional modules",
            "Module Collection summarizes compatible work-level values. Module Work Results, Module Structure, Module Coverage, and Module Provenance retain their scopes, denominators, configurations, and source evidence.",
        ),
        (
            "Lexical diversity aggregation",
            "Equal-work summaries and ordered pooled-token calculations are separate. VerseVAD does not average MATTR, HD-D, or MTLD as though work-level values were interchangeable pooled observations.",
        ),
        (
            "Emotion and sentiment",
            "Eight emotion associations, positive/negative sentiment, and supplied emotion intensity remain separately labeled constructs.",
        ),
        (
            "Token-weighted volume profile",
            "Pools included matched observations. Long works contribute more because they contain more of the volume's words.",
        ),
        (
            "Work-weighted volume profile",
            "Averages eligible work-level token means. Every poem contributes one score regardless of length.",
        ),
        (
            "Within-work token weighting",
            "Every included occurrence contributes, so repetitions matter.",
        ),
        (
            "Within-work type weighting",
            "Every distinct matched lexicon entry contributes once; this describes vocabulary breadth.",
        ),
        (
            "Stopword sensitivity",
            "Every VAD result is reported as all matched tokens and, when enabled, with active stopwords excluded. Neither is labeled more accurate.",
        ),
        (
            "Cumulative lexical load",
            "Length-sensitive normalized rating totals and midpoint deviations. These are lexical evidence, not measured reader response.",
        ),
        (
            "Missing data",
            "Unmatched words and missing work scores remain missing. VerseVAD never assigns them a neutral value.",
        ),
        (
            "Interpretive scope",
            "Results describe normative ratings and lexical associations under a matching policy—not a poem, speaker, author, or reader's emotion.",
        ),
    )
    for row_number, (term, explanation) in enumerate(guidance, start=5):
        readme.cell(row_number, 1, term).font = Font(bold=True, color=INK)
        readme.cell(row_number, 2, explanation).alignment = Alignment(wrap_text=True, vertical="top")
        if row_number % 2:
            readme.cell(row_number, 1).fill = PatternFill("solid", fgColor=PALE)
            readme.cell(row_number, 2).fill = PatternFill("solid", fgColor=PALE)
    readme.column_dimensions["A"].width = 34
    readme.column_dimensions["B"].width = 100
    readme.freeze_panes = "A5"

    profiles = corpus_vad_profiles(
        metrics,
        total_works=len({row.text_id for row in metrics}),
    )
    profile_headers = (
        "Lexicon",
        "Analysis view",
        "Dimension",
        "Works included",
        "Works omitted",
        "Matched observations",
        "Lexical tokens",
        "Volume coverage",
        "Token-weighted volume mean",
        "Work-weighted volume mean",
        "Work minus token difference",
    )
    profile_rows = (
        (
            row.lexicon,
            _label(row.analysis_view),
            row.dimension.title(),
            row.works_included,
            row.works_omitted,
            row.matched_observations,
            row.lexical_tokens,
            row.volume_coverage,
            row.token_weighted_volume_mean,
            row.work_weighted_volume_mean,
            row.work_minus_token_difference,
        )
        for row in profiles
    )
    _write_sheet(
        workbook,
        title="Corpus Profiles",
        purpose=(
            "Two collection views: pooled matched observations versus equal weight per eligible work. "
            "Their difference is analytically meaningful."
        ),
        headers=profile_headers,
        rows=profile_rows,
        table_name="CorpusProfiles",
    )
    profile_sheet = workbook["Corpus Profiles"]
    if profiles:
        chart = BarChart()
        chart.type = "bar"
        chart.style = 10
        chart.title = "Token- and work-weighted collection means"
        chart.x_axis.title = "Normalized mean (0–1)"
        chart.x_axis.scaling.min = 0
        chart.x_axis.scaling.max = 1
        chart.x_axis.majorUnit = 0.2
        chart.x_axis.numFmt = "0.0"
        chart.height = 7
        chart.width = 16
        data = Reference(profile_sheet, min_col=9, max_col=10, min_row=4, max_row=4 + len(profiles))
        categories = Reference(profile_sheet, min_col=3, min_row=5, max_row=4 + len(profiles))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.series[0].tx = SeriesLabel(v="Token-weighted volume mean")
        chart.series[0].graphicalProperties.solidFill = "4F81BD"
        chart.series[1].tx = SeriesLabel(v="Work-weighted volume mean")
        chart.series[1].graphicalProperties.solidFill = "C0504D"
        chart.legend.position = "b"
        profile_sheet.add_chart(chart, "L4")

    module_profiles = corpus_module_profiles(
        module_metrics,
        total_works=len({row.text_id for row in module_metrics}),
    )
    _write_sheet(
        workbook,
        title="Module Collection",
        purpose=(
            "Compatible numeric document metrics summarized across works. "
            "Equal-work and observation-weighted means remain separate; explicitly "
            "pooled calculations are labeled by their aggregation method."
        ),
        headers=(
            "Module",
            "Metric",
            "Unit",
            "Weighting",
            "Works included",
            "Works omitted",
            "Equal-work mean",
            "Observation-weighted mean",
            "Observations",
            "Configuration",
            "Aggregation method",
            "Note",
        ),
        rows=tuple(
            (
                row.module_name,
                row.metric_id,
                row.unit,
                row.weighting,
                row.works_included,
                row.works_omitted,
                row.equal_work_mean,
                row.observation_weighted_mean,
                row.total_observations or None,
                row.configuration_id,
                "compatible_work_values",
                row.note,
            )
            for row in module_profiles
        )
        + tuple(
            (
                row.module_name,
                row.metric_id,
                row.unit,
                "",
                row.works_included,
                row.works_omitted,
                _cell_value(row.value),
                None,
                row.observation_count,
                row.configuration_id,
                row.aggregation_method,
                row.note,
            )
            for row in module_aggregates
        ),
        table_name="ModuleCollection",
    )
    module_categories = corpus_module_category_profiles(module_metrics)
    _write_sheet(
        workbook,
        title="Module Categories",
        purpose=(
            "Work-level prevalence of selected meter and rhyme categories. "
            "These distributions do not declare one corpus-wide scheme."
        ),
        headers=(
            "Module",
            "Metric",
            "Category",
            "Works with category",
            "Eligible works",
            "Prevalence",
            "Configuration",
            "Note",
        ),
        rows=(
            (
                row.module_name,
                row.metric_id,
                row.category,
                row.works_with_category,
                row.works_included,
                row.prevalence,
                row.configuration_id,
                row.note,
            )
            for row in module_categories
        ),
        table_name="ModuleCategories",
    )

    module_headers = (
        "Work",
        "Author",
        "Collection",
        "Date",
        "Genre",
        "Module",
        "Module version",
        "Metric",
        "Value",
        "Layer",
        "Scope",
        "Scope ID",
        "Unit",
        "Weighting",
        "Denominator",
        "Observations",
        "Configuration",
        "Note",
    )
    for title, purpose, rows, table_name in (
        (
            "Module Work Results",
            "Document-scope results from every enabled non-affective module.",
            (row for row in module_metrics if row.scope == "document"),
            "ModuleWorkResults",
        ),
        (
            "Module Structure",
            "Line-, stanza-, token-, type-, distribution-, and other non-document module results retain their original scope IDs.",
            (row for row in module_metrics if row.scope != "document"),
            "ModuleStructure",
        ),
    ):
        _write_sheet(
            workbook,
            title=title,
            purpose=purpose,
            headers=module_headers,
            rows=(
                (
                    row.title,
                    row.author,
                    row.collection,
                    row.date_label,
                    row.genre,
                    row.module_name,
                    row.module_version,
                    row.metric_id,
                    _cell_value(row.value),
                    row.layer,
                    row.scope,
                    row.scope_id,
                    row.unit,
                    row.weighting,
                    row.denominator,
                    row.observation_count,
                    row.configuration_id,
                    row.note,
                )
                for row in rows
            ),
            table_name=table_name,
        )

    _write_sheet(
        workbook,
        title="Module Coverage",
        purpose=(
            "Eligible, matched, and unmatched module evidence. Empty coverage rates "
            "remain missing rather than becoming zero or neutral."
        ),
        headers=(
            "Work",
            "Module",
            "Coverage measure",
            "Scope",
            "Scope ID",
            "Eligible",
            "Matched",
            "Unmatched",
            "Coverage",
            "Unit",
            "Unmatched items",
            "Configuration",
            "Note",
        ),
        rows=(
            (
                row.title,
                row.module_name,
                row.coverage_id,
                row.scope,
                row.scope_id,
                row.eligible_count,
                row.matched_count,
                row.unmatched_count,
                row.coverage_rate,
                row.unit,
                ", ".join(row.unmatched_items),
                row.configuration_id,
                row.note,
            )
            for row in module_coverage
        ),
        table_name="ModuleCoverage",
    )
    _write_sheet(
        workbook,
        title="Module Provenance",
        purpose=(
            "One row per persisted module result with stable IDs, configuration, "
            "source-text hash, and serialized local provenance."
        ),
        headers=(
            "Work",
            "Module",
            "Module version",
            "Result ID",
            "Run ID",
            "Text ID",
            "Text version ID",
            "Configuration",
            "Scenario",
            "Source text SHA-256",
            "Completed at",
            "Provenance",
        ),
        rows=(
            (
                row.title,
                row.module_name,
                row.module_version,
                row.result_id,
                row.run_id,
                row.text_id,
                row.text_version_id,
                row.configuration_id,
                row.scenario_id,
                row.source_text_sha256,
                row.completed_at,
                _cell_value(dict(row.provenance)),
            )
            for row in module_results
        ),
        table_name="ModuleProvenance",
    )
    _write_sheet(
        workbook,
        title="Module Warnings",
        purpose=(
            "Plain-language module warnings remain linked to their work and exact "
            "configuration. Technical detail is retained for diagnosis."
        ),
        headers=(
            "Work",
            "Module",
            "Severity",
            "Code",
            "Message",
            "Technical detail",
            "Configuration",
            "Run ID",
            "Completed at",
        ),
        rows=(
            (
                row.title,
                row.module_name,
                row.severity,
                row.code,
                row.message,
                row.technical_detail,
                row.configuration_id,
                row.run_id,
                row.completed_at,
            )
            for row in module_warnings
        ),
        table_name="ModuleWarnings",
    )

    work_vad = [
        row
        for row in metrics
        if row.metric in {"vad_mean", "vad_standard_deviation"}
        and row.scale == "normalized_0_1"
    ]
    _write_sheet(
        workbook,
        title="Work VAD",
        purpose="One row per work, lexicon, dimension, and token/type weighting on the derived 0–1 scale.",
        headers=(
            "Work",
            "Author",
            "Collection",
            "Date",
            "Genre",
            "Lexicon",
            "Analysis view",
            "Dimension",
            "Weighting",
            "Statistic",
            "Value",
            "Observations",
            "Lexical tokens",
            "Coverage",
        ),
        rows=(
            (
                row.title,
                row.author,
                row.collection,
                row.date_label,
                row.genre,
                row.lexicon,
                _label(row.analysis_view),
                row.dimension.title(),
                row.weighting.title(),
                "Mean" if row.metric == "vad_mean" else "Population standard deviation",
                row.value,
                row.observations,
                row.lexical_tokens,
                row.coverage,
            )
            for row in work_vad
        ),
        table_name="WorkVad",
    )

    cumulative_metric_names = {
        "vad_rating_total",
        "vad_above_midpoint_load",
        "vad_below_midpoint_load",
        "vad_net_midpoint_load",
        "vad_absolute_midpoint_load",
    }
    load_metrics = [row for row in metrics if row.metric in cumulative_metric_names]
    _write_sheet(
        workbook,
        title="Cumulative Load",
        purpose=(
            "Length-sensitive sums by work. Above/below/absolute loads use distance from the 0.5 midpoint; "
            "net load permits directional cancellation."
        ),
        headers=(
            "Work",
            "Collection",
            "Lexicon",
            "Analysis view",
            "Dimension",
            "Measure",
            "Value",
            "Matched observations",
            "Lexical tokens",
            "Coverage",
        ),
        rows=(
            (
                row.title,
                row.collection,
                row.lexicon,
                _label(row.analysis_view),
                row.dimension.title(),
                _label(row.metric),
                row.value,
                row.observations,
                row.lexical_tokens,
                row.coverage,
            )
            for row in load_metrics
        ),
        table_name="CumulativeLoad",
    )

    other = [row for row in metrics if not row.metric.startswith("vad_")]
    _write_sheet(
        workbook,
        title="Coverage and Emotion",
        purpose="Coverage, categorical association, and supplied emotion-intensity metrics remain separate constructs.",
        headers=(
            "Work",
            "Collection",
            "Lexicon",
            "Construct",
            "Analysis view",
            "Metric",
            "Category",
            "Weighting",
            "Scale",
            "Value",
            "Observations",
            "Matched tokens",
            "Lexical tokens",
            "Coverage",
            "Denominator",
        ),
        rows=(
            (
                row.title,
                row.collection,
                row.lexicon,
                _construct_label(row.metric, row.category),
                _label(row.analysis_view),
                _label(row.metric),
                row.category.title(),
                row.weighting.title(),
                row.scale,
                row.value,
                row.observations,
                row.matched_tokens,
                row.lexical_tokens,
                row.coverage,
                row.denominator,
            )
            for row in other
        ),
        table_name="CoverageEmotion",
    )

    if part_of_speech_rows:
        _write_sheet(
            workbook,
            title="Part of Speech",
            purpose=(
                "Model-assigned universal part-of-speech counts and relative shares "
                "use all eligible lexical tokens, independently of lexicon coverage. "
                "Noun combines source NOUN and PROPN tags; Verb combines VERB "
                "and AUX tags."
            ),
            headers=(
                "Scope",
                "Profile Level",
                "Work",
                "Collection",
                "Source POS tag(s)",
                "Part of speech",
                "Token count",
                "Share of lexical tokens",
                "Unique normalized types",
                "Examples",
                "Lexical-token denominator",
                "Model",
            ),
            rows=(
                (
                    row.get("Scope", ""),
                    row.get("Profile Level", ""),
                    row.get("Work", ""),
                    row.get("Collection", ""),
                    row.get("Source POS tag(s)", ""),
                    row.get("Part of speech", ""),
                    row.get("Token count", ""),
                    row.get("Share of lexical tokens", ""),
                    row.get("Unique normalized types", ""),
                    row.get("Examples", ""),
                    row.get("Lexical-token denominator", ""),
                    row.get("Model", ""),
                )
                for row in part_of_speech_rows
            ),
            table_name="PartOfSpeech",
        )

    _write_sheet(
        workbook,
        title="Unmatched QC",
        purpose="Unmatched vocabulary from the latest complete corpus batch, joined to persistent local review notes.",
        headers=(
            "Work",
            "Lexicon",
            "Surface form",
            "Normalized form",
            "Frequency",
            "POS",
            "Proposed lemma",
            "Example line",
            "Example context",
            "Status",
            "Research note",
            "Proposed mapping",
        ),
        rows=(
            (
                row.text_title,
                row.lexicon,
                row.display_form,
                row.normalized_form,
                row.frequency,
                row.pos,
                row.proposed_lemma,
                row.example_line,
                row.example_context,
                row.status,
                row.note,
                row.proposed_mapping,
            )
            for row in unmatched
        ),
        table_name="UnmatchedQc",
    )

    if review_decisions:
        _write_sheet(
            workbook,
            title="Review Decisions",
            purpose=(
                "Exact active decision revisions pinned to the exported scenario. "
                "Flags do not change scores; exclusions and mappings apply only "
                "within this recorded scenario version."
            ),
            headers=(
                "Decision ID",
                "Decision revision ID",
                "Action",
                "Scope",
                "Lexicon ID",
                "Source form",
                "Mapping target",
                "Project ID",
                "Text ID",
                "Text version ID",
                "Token position",
                "Risk category",
                "Rationale",
            ),
            rows=(
                (
                    decision.get("decision_id", ""),
                    decision.get("decision_revision_id", ""),
                    _label(str(decision.get("action", ""))),
                    _label(str(decision.get("scope", ""))),
                    decision.get("lexicon_id", ""),
                    decision.get("source_form", ""),
                    decision.get("mapping_target", ""),
                    decision.get("project_id", ""),
                    decision.get("text_id", ""),
                    decision.get("text_version_id", ""),
                    decision.get("token_position", ""),
                    _label(str(decision.get("risk_category", ""))),
                    decision.get("rationale", ""),
                )
                for decision in review_decisions
            ),
            table_name="ReviewDecisions",
        )

    metadata_headers = (
        "Text ID",
        "Text version ID",
        "Title",
        "Author",
        "Collection",
        "Date",
        "Genre",
        "Source path",
        "SHA-256",
        "Imported at",
        "Notes",
        "Custom metadata",
    )
    _write_sheet(
        workbook,
        title="Text Metadata",
        purpose="Stable identities and source hashes for the active preserved text versions. Original text is not duplicated in this workbook.",
        headers=metadata_headers,
        rows=(
            (
                text.text_id,
                text.text_version_id,
                text.title,
                text.author,
                text.collection,
                text.date_label,
                text.genre,
                text.relative_path,
                text.text_sha256,
                text.imported_at,
                text.notes,
                str(dict(text.custom_metadata)),
            )
            for text in texts
        ),
        table_name="TextMetadata",
    )

    if methodology:
        stopword = methodology.get("stopword_policy") or {}
        if isinstance(stopword, Mapping):
            methodology_rows = (
                ("Software version", methodology.get("software_version", "")),
                ("Scenario", methodology.get("scenario_id", "")),
                (
                    "Scenario version",
                    methodology.get("scenario_version_id", ""),
                ),
                (
                    "Active review decision revisions",
                    len(methodology.get("review_decisions", ())),
                ),
                ("Phrase policy", methodology.get("phrase_policy", "")),
                (
                    "Minimum match requirement",
                    methodology.get("minimum_match_requirement", ""),
                ),
                (
                    "Optional modules",
                    json.dumps(
                        methodology.get("optional_modules", ()),
                        ensure_ascii=False,
                        sort_keys=True,
                    ),
                ),
                (
                    "Optional module configurations",
                    json.dumps(
                        methodology.get(
                            "optional_module_configurations",
                            {},
                        ),
                        ensure_ascii=False,
                        sort_keys=True,
                    ),
                ),
                ("Stopword mode", stopword.get("mode", "")),
                ("Stopword source", stopword.get("source", "")),
                ("Stopword library version", stopword.get("library_version", "")),
                ("Stopword list version", stopword.get("list_version", "")),
                ("Standard stopword count", stopword.get("standard_word_count", "")),
                ("Standard list SHA-256", stopword.get("standard_list_sha256", "")),
                ("Active list SHA-256", stopword.get("active_list_sha256", "")),
                (
                    "Active stopwords",
                    ", ".join(stopword.get("active_words", ())),
                ),
                (
                    "Protected words",
                    ", ".join(stopword.get("protected_words", ())),
                ),
                (
                    "Custom additions",
                    ", ".join(stopword.get("custom_additions", ())),
                ),
                (
                    "Custom removals",
                    ", ".join(stopword.get("custom_removals", ())),
                ),
            )
            _write_sheet(
                workbook,
                title="Methodology",
                purpose=(
                    "Reproducibility settings recorded with the latest complete "
                    "analysis batch."
                ),
                headers=("Field", "Value"),
                rows=methodology_rows,
                table_name="Methodology",
            )

    workbook.properties.title = f"VerseVAD corpus analysis — {project.title}"
    workbook.properties.subject = "Local descriptive affective lexicon analysis"
    workbook.properties.creator = "VerseVAD"
    workbook.calculation.fullCalcOnLoad = True
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()
