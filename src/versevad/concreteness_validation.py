"""Invented, hand-calculated validation for the concreteness module."""

from __future__ import annotations

import hashlib
import math
import tempfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook

from versevad.core import ModuleInput, ResourceSpec
from versevad.lexical_semantic.concreteness import (
    ConcretenessConfiguration,
    ConcretenessMatchMethod,
    ConcretenessModule,
)
from versevad.preprocessing import SpacyEnglishPreprocessor, create_text_document


_HEADER = (
    "Word",
    "Bigram",
    "Conc.M",
    "Conc.SD",
    "Unknown",
    "Total",
    "Percent_known",
    "SUBTLEX",
)
_ROWS = (
    ("dark", 0, 1.5, 0.5, 0, 30, 1.0, 50),
    ("night", 0, 3.0, 0.5, 0, 30, 1.0, 50),
    ("dark night", 1, 4.5, 0.5, 0, 30, 1.0, 5),
    ("stone", 0, 5.0, 0.2, 0, 30, 1.0, 100),
    ("idea", 0, 1.0, 0.4, 0, 30, 1.0, 100),
)
_TEXT = "dark night\nstone stones\nidea quorvax"


@dataclass(frozen=True)
class SyntheticConcretenessValidation:
    """Compact result used by the command-line check and its regression test."""

    eligible_tokens: int
    rated_tokens: int
    token_coverage: float | None
    mean_normative_concreteness: float | None
    matched_expression_occurrences: int
    phrase_occurrences: int
    exact_phrase_tokens: int
    exact_tokens: int
    lemma_tokens: int
    unmatched_tokens: int
    source_unchanged: bool


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_fixture(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(_HEADER)
    for row in _ROWS:
        sheet.append(row)
    workbook.create_sheet("Sheet2")
    workbook.create_sheet("Sheet3")
    workbook.save(path)


def run_synthetic_concreteness_validation(
) -> tuple[SyntheticConcretenessValidation, tuple[str, ...]]:
    """Run an invented example with values that can be checked by hand."""

    with tempfile.TemporaryDirectory(prefix="versevad-concreteness-") as directory:
        root = Path(directory)
        source = root / "synthetic_concreteness.xlsx"
        _write_fixture(source)
        before = _sha256(source)
        resource = ResourceSpec(
            resource_id="synthetic-concreteness-validation",
            display_name="Synthetic concreteness validation fixture",
            relative_path=source.name,
            version="synthetic-v1",
            accepted_sha256=(before,),
            citation="Invented VerseVAD validation data.",
            license_notice="Synthetic data generated locally for validation.",
        )
        module = ConcretenessModule(root, resource_spec=resource)
        poem = SpacyEnglishPreprocessor().process_document(
            create_text_document(
                "concreteness-validation",
                "Invented concreteness validation",
                _TEXT,
            )
        )
        result = module.analyze_detailed(
            ModuleInput.from_poem_document(poem),
            ConcretenessConfiguration(exclude_proper_nouns=False),
        )
        after = _sha256(source)

    method_counts = {
        method: sum(row.match_method is method for row in result.token_audit)
        for method in ConcretenessMatchMethod
    }
    report = SyntheticConcretenessValidation(
        eligible_tokens=result.summary.eligible_token_count,
        rated_tokens=result.summary.rated_token_count,
        token_coverage=result.summary.token_coverage,
        mean_normative_concreteness=result.summary.statistics.mean,
        matched_expression_occurrences=(
            result.summary.matched_expression_occurrence_count
        ),
        phrase_occurrences=len(
            {
                row.match_group_id
                for row in result.token_audit
                if row.match_method is ConcretenessMatchMethod.EXACT_PHRASE
            }
        ),
        exact_phrase_tokens=method_counts[ConcretenessMatchMethod.EXACT_PHRASE],
        exact_tokens=method_counts[ConcretenessMatchMethod.EXACT],
        lemma_tokens=method_counts[ConcretenessMatchMethod.LEMMA],
        unmatched_tokens=method_counts[ConcretenessMatchMethod.UNMATCHED],
        source_unchanged=before == after,
    )

    problems: list[str] = []
    expected = {
        "eligible_tokens": 6,
        "rated_tokens": 5,
        "matched_expression_occurrences": 4,
        "phrase_occurrences": 1,
        "exact_phrase_tokens": 2,
        "exact_tokens": 2,
        "lemma_tokens": 1,
        "unmatched_tokens": 1,
    }
    for field, value in expected.items():
        if getattr(report, field) != value:
            problems.append(
                f"{field} was {getattr(report, field)!r}; expected {value!r}."
            )
    if report.token_coverage is None or not math.isclose(
        report.token_coverage,
        5 / 6,
    ):
        problems.append(
            "Token coverage did not equal the hand-calculated 5/6."
        )
    if report.mean_normative_concreteness is None or not math.isclose(
        report.mean_normative_concreteness,
        4.0,
    ):
        problems.append(
            "Mean normative concreteness did not equal the hand-calculated 4.0."
        )
    if not report.source_unchanged:
        problems.append("The synthetic source workbook changed during analysis.")
    if any(
        row.rating is not None
        for row in result.token_audit
        if not row.included
    ):
        problems.append("An unmatched or ineligible token received a rating.")

    return report, tuple(problems)


def main() -> int:
    report, problems = run_synthetic_concreteness_validation()
    if problems:
        print("VerseVAD's concreteness validation did not match expectations.")
        for problem in problems:
            print(f"- {problem}")
        return 1

    print("VerseVAD concreteness validation passed.")
    print(
        "Rated lexical tokens: "
        f"{report.rated_tokens}/{report.eligible_tokens} "
        f"({report.token_coverage:.1%} coverage)."
    )
    print(
        "Mean normative lexical concreteness of matched tokens: "
        f"{report.mean_normative_concreteness:.6f} on the 1-5 source scale."
    )
    print(
        "The exact two-word expression, exact forms, lemma fallback, and "
        "unmatched token all followed the expected audit path."
    )
    print("The generated synthetic workbook remained unchanged.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
