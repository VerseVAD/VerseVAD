"""Invented, hand-calculated validation for the optional Kuperman AoA module."""

from __future__ import annotations

import hashlib
import math
import tempfile
from dataclasses import dataclass, replace
from pathlib import Path

from openpyxl import Workbook

from versevad.adapters.kuperman_aoa import REQUIRED_COLUMNS
from versevad.core import ModuleInput, ResourceSpec
from versevad.lexical_semantic.aoa import (
    AoAConfiguration,
    AoAMatchMethod,
    AoAModule,
)
from versevad.preprocessing import SpacyEnglishPreprocessor, create_text_document


def _row(
    term: str,
    mean_age: float,
    *,
    numeric_responses: int = 20,
) -> tuple[object, ...]:
    total = 20
    return (
        term,
        total,
        numeric_responses,
        10.0,
        mean_age,
        1.5,
        numeric_responses / total,
    )


_ROWS = (
    _row("early", 3.0, numeric_responses=4),
    _row("middle", 8.0),
    _row("later", 14.0),
    _row("the", 4.0),
    _row("stone", 5.0),
    _row("runs", 6.0),
    _row("bright", 7.0),
    _row("swiftly", 8.0),
    _row("and", 4.5),
    _row("she", 4.0),
    _row("can", 4.5),
    _row("under", 6.0),
)
_TEXT = "early early\nmiddle middles\nlater quorvax"
_CONTENT_TEXT = "the stone runs bright swiftly and she can under"


@dataclass(frozen=True)
class SyntheticAoAValidation:
    eligible_tokens: int
    matched_tokens: int
    token_coverage: float | None
    mean_aoa: float | None
    median_aoa: float | None
    early_tokens: int
    middle_tokens: int
    later_tokens: int
    exact_tokens: int
    lemma_tokens: int
    unmatched_tokens: int
    low_response_tokens: int
    content_scope_eligible_tokens: int
    content_scope_excluded_tokens: int
    source_unchanged: bool


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _write_fixture(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(REQUIRED_COLUMNS)
    for row in _ROWS:
        sheet.append(row)
    workbook.create_sheet("Sheet2")
    workbook.create_sheet("Sheet3")
    workbook.save(path)


def run_synthetic_aoa_validation(
) -> tuple[SyntheticAoAValidation, tuple[str, ...]]:
    """Run invented examples whose expected AoA results can be checked by hand."""

    with tempfile.TemporaryDirectory(prefix="versevad-aoa-") as directory:
        root = Path(directory)
        source = root / "synthetic_kuperman_aoa.xlsx"
        _write_fixture(source)
        before = _sha256(source)
        resource = ResourceSpec(
            resource_id="synthetic-aoa-validation",
            display_name="Synthetic AoA validation fixture",
            relative_path=source.name,
            version="synthetic-v1",
            accepted_sha256=(before,),
            citation="Invented VerseVAD validation data.",
            license_notice="Synthetic data generated locally for validation.",
        )
        module = AoAModule(root, resource_spec=resource)
        processor = SpacyEnglishPreprocessor()
        poem = processor.process_document(
            create_text_document(
                "aoa-validation",
                "Invented AoA validation",
                _TEXT,
            )
        )
        result = module.analyze_detailed(
            ModuleInput.from_poem_document(poem),
            AoAConfiguration(exclude_proper_nouns=False),
        )

        content_poem = processor.process_document(
            create_text_document(
                "aoa-content-validation",
                "Invented AoA content-scope validation",
                _CONTENT_TEXT,
            )
        )
        tags = {
            "the": "DET",
            "stone": "NOUN",
            "runs": "VERB",
            "bright": "ADJ",
            "swiftly": "ADV",
            "and": "CCONJ",
            "she": "PRON",
            "can": "AUX",
            "under": "ADP",
        }
        content_tokens = tuple(
            replace(token, part_of_speech=tags[token.normalized_form])
            for token in content_poem.tokens
        )
        content_result = module.analyze_detailed(
            ModuleInput(
                document=content_poem.source,
                tokens=content_tokens,
                preprocessing=content_poem.preprocessing,
            ),
            AoAConfiguration(
                exclude_proper_nouns=False,
                content_words_only=True,
            ),
        )
        after = _sha256(source)

    method_counts = {
        method: sum(row.match_method is method for row in result.token_audit)
        for method in AoAMatchMethod
    }
    bands = {band.band_id: band for band in result.summary.bands}
    report = SyntheticAoAValidation(
        eligible_tokens=result.summary.eligible_token_count,
        matched_tokens=result.summary.matched_token_count,
        token_coverage=result.summary.token_coverage,
        mean_aoa=result.summary.statistics.mean,
        median_aoa=result.summary.statistics.median,
        early_tokens=bands["early_acquired"].token_count,
        middle_tokens=bands["middle_range"].token_count,
        later_tokens=bands["later_acquired"].token_count,
        exact_tokens=method_counts[AoAMatchMethod.EXACT],
        lemma_tokens=method_counts[AoAMatchMethod.LEMMA],
        unmatched_tokens=method_counts[AoAMatchMethod.UNMATCHED],
        low_response_tokens=result.summary.low_response_token_count,
        content_scope_eligible_tokens=(
            content_result.summary.eligible_token_count
        ),
        content_scope_excluded_tokens=sum(
            not row.eligible for row in content_result.token_audit
        ),
        source_unchanged=before == after,
    )

    problems = []
    expected = {
        "eligible_tokens": 6,
        "matched_tokens": 5,
        "early_tokens": 2,
        "middle_tokens": 2,
        "later_tokens": 1,
        "exact_tokens": 4,
        "lemma_tokens": 1,
        "unmatched_tokens": 1,
        "low_response_tokens": 2,
        "content_scope_eligible_tokens": 4,
        "content_scope_excluded_tokens": 5,
    }
    for field, value in expected.items():
        if getattr(report, field) != value:
            problems.append(
                f"{field} was {getattr(report, field)!r}; expected {value!r}."
            )
    if report.token_coverage is None or not math.isclose(
        report.token_coverage, 5 / 6
    ):
        problems.append("Token coverage did not equal the hand-calculated 5/6.")
    if report.mean_aoa is None or not math.isclose(report.mean_aoa, 7.2):
        problems.append("Mean AoA did not equal the hand-calculated 7.2.")
    if report.median_aoa is None or not math.isclose(report.median_aoa, 8.0):
        problems.append("Median AoA did not equal the hand-calculated 8.0.")
    if not report.source_unchanged:
        problems.append("The synthetic source workbook changed during analysis.")
    if any(
        row.mean_age is not None
        for row in result.token_audit
        if not row.included
    ):
        problems.append("An unmatched or ineligible token received a mean age.")
    return report, tuple(problems)


def main() -> int:
    report, problems = run_synthetic_aoa_validation()
    if problems:
        print("VerseVAD's AoA validation did not match expectations.")
        for problem in problems:
            print(f"- {problem}")
        return 1

    print("VerseVAD Kuperman age-of-acquisition validation passed.")
    print(
        "Matched lexical tokens: "
        f"{report.matched_tokens}/{report.eligible_tokens} "
        f"({report.token_coverage:.1%} coverage)."
    )
    print(
        "Mean/median normative source age of matched tokens: "
        f"{report.mean_aoa:.6f}/{report.median_aoa:.6f} years."
    )
    print(
        "Exact forms, the explicit lemma fallback, unmatched values, source "
        "response evidence, bands, and optional contextual "
        "NOUN/VERB/ADJ/ADV-only scope followed the expected audit."
    )
    print("The generated synthetic workbook remained unchanged.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
