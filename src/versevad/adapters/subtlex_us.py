"""Read-only adapter for the official SUBTLEX-US Zipf workbook."""

from __future__ import annotations

import hashlib
import math
from dataclasses import dataclass
from pathlib import Path
from types import MappingProxyType
from typing import Mapping

from openpyxl import load_workbook

from versevad.normalization import normalize_lookup


REQUIRED_COLUMNS = (
    "Word",
    "FREQcount",
    "CDcount",
    "FREQlow",
    "Cdlow",
    "SUBTLWF",
    "Lg10WF",
    "SUBTLCD",
    "Lg10CD",
    "Dom_PoS_SUBTLEX",
    "Freq_dom_PoS_SUBTLEX",
    "Percentage_dom_PoS",
    "All_PoS_SUBTLEX",
    "All_freqs_SUBTLEX",
    "Zipf-value",
)


class SubtlexUsAdapterError(RuntimeError):
    """Plain-language adapter failure that confirms the source stayed unchanged."""

    def __init__(self, message: str, technical_detail: str = "") -> None:
        super().__init__(message)
        self.technical_detail = technical_detail
        self.data_changed = False


@dataclass(frozen=True)
class SubtlexUsEntry:
    source_term: str
    lookup_form: str
    source_row: int
    frequency_count: int
    contextual_diversity_count: int
    lowercase_frequency_count: int
    lowercase_contextual_diversity_count: int
    frequency_per_million: float
    log10_frequency: float
    contextual_diversity_percent: float
    log10_contextual_diversity: float
    dominant_source_pos: str
    dominant_source_pos_frequency: int | None
    dominant_source_pos_proportion: float | None
    all_source_pos: str
    all_source_pos_frequencies: str
    zipf_value: float


@dataclass(frozen=True)
class SubtlexUsValidation:
    source_path: Path
    source_sha256: str
    total_rows: int
    usable_entries: int
    blank_terms: int
    malformed_rows: int
    duplicate_keys: int
    out_of_range_values: int
    errors: tuple[str, ...] = ()
    warnings: tuple[str, ...] = ()

    @property
    def is_valid(self) -> bool:
        return not self.errors


@dataclass(frozen=True)
class SubtlexUsLexicon:
    entries: Mapping[str, SubtlexUsEntry]
    validation: SubtlexUsValidation

    @classmethod
    def create(
        cls,
        entries: Mapping[str, SubtlexUsEntry],
        validation: SubtlexUsValidation,
    ) -> SubtlexUsLexicon:
        return cls(MappingProxyType(dict(entries)), validation)

    def lookup(self, lookup_form: str) -> SubtlexUsEntry | None:
        return self.entries.get(lookup_form)


class SubtlexUsAdapter:
    """Parse the pinned official workbook without saving or rewriting it."""

    adapter_version = "1.0.0"

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _finite_number(value: object, label: str) -> float:
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise ValueError(f"{label} is not numeric")
        number = float(value)
        if not math.isfinite(number):
            raise ValueError(f"{label} is not finite")
        return number

    @classmethod
    def _integer(cls, value: object, label: str) -> int:
        number = cls._finite_number(value, label)
        if not number.is_integer():
            raise ValueError(f"{label} is not an integer")
        return int(number)

    @staticmethod
    def _text(value: object, label: str) -> str:
        if not isinstance(value, str) or not value.strip():
            raise ValueError(f"{label} is blank or non-text")
        return value

    @classmethod
    def _optional_integer(cls, value: object, label: str) -> int | None:
        if value == "#N/A":
            return None
        return cls._integer(value, label)

    @classmethod
    def _optional_finite_number(
        cls, value: object, label: str
    ) -> float | None:
        if value == "#N/A":
            return None
        return cls._finite_number(value, label)

    def load(self, source_path: Path | str) -> SubtlexUsLexicon:
        path = Path(source_path)
        if not path.is_file():
            raise SubtlexUsAdapterError(
                "The SUBTLEX-US Zipf workbook was not found.",
                f"Expected a readable local file at {path}.",
            )
        try:
            source_sha256 = self._sha256(path)
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as error:
            raise SubtlexUsAdapterError(
                "The SUBTLEX-US Zipf workbook could not be opened.",
                f"{type(error).__name__}: {error}",
            ) from error

        try:
            if "out1g" not in workbook.sheetnames:
                raise SubtlexUsAdapterError(
                    "The SUBTLEX-US workbook does not contain the expected data sheet.",
                    "Expected a worksheet named out1g.",
                )
            sheet = workbook["out1g"]
            rows = sheet.iter_rows(values_only=True)
            try:
                header = tuple(next(rows))
            except StopIteration as error:
                raise SubtlexUsAdapterError(
                    "The SUBTLEX-US workbook is empty.",
                    "The out1g sheet contains no header row.",
                ) from error
            if header != REQUIRED_COLUMNS:
                missing = [
                    column for column in REQUIRED_COLUMNS if column not in header
                ]
                raise SubtlexUsAdapterError(
                    "The SUBTLEX-US workbook does not have the expected columns.",
                    (
                        f"Expected exactly {REQUIRED_COLUMNS!r}; found {header!r}. "
                        f"Missing: {missing!r}."
                    ),
                )

            entries: dict[str, SubtlexUsEntry] = {}
            total_rows = 0
            blank_terms = 0
            malformed_rows = 0
            duplicate_keys = 0
            out_of_range_values = 0
            problems: list[str] = []

            for source_row, row in enumerate(rows, start=2):
                if not row or all(value is None for value in row):
                    continue
                total_rows += 1
                if len(row) != len(REQUIRED_COLUMNS):
                    malformed_rows += 1
                    problems.append(
                        f"row {source_row}: expected 15 cells, found {len(row)}"
                    )
                    continue
                term_value = row[0]
                if not isinstance(term_value, str) or not term_value.strip():
                    blank_terms += 1
                    problems.append(f"row {source_row}: blank or non-text Word value")
                    continue
                source_term = term_value
                lookup_form = normalize_lookup(source_term)
                try:
                    frequency_count = self._integer(row[1], "FREQcount")
                    contextual_diversity_count = self._integer(row[2], "CDcount")
                    lowercase_frequency_count = self._integer(row[3], "FREQlow")
                    lowercase_contextual_diversity_count = self._integer(
                        row[4], "Cdlow"
                    )
                    frequency_per_million = self._finite_number(row[5], "SUBTLWF")
                    log10_frequency = self._finite_number(row[6], "Lg10WF")
                    contextual_diversity_percent = self._finite_number(
                        row[7], "SUBTLCD"
                    )
                    log10_contextual_diversity = self._finite_number(
                        row[8], "Lg10CD"
                    )
                    dominant_source_pos = self._text(row[9], "Dom_PoS_SUBTLEX")
                    dominant_source_pos_frequency = self._optional_integer(
                        row[10], "Freq_dom_PoS_SUBTLEX"
                    )
                    dominant_source_pos_proportion = self._optional_finite_number(
                        row[11], "Percentage_dom_PoS"
                    )
                    all_source_pos = self._text(row[12], "All_PoS_SUBTLEX")
                    all_source_pos_frequencies = str(row[13])
                    if row[13] is None or not all_source_pos_frequencies.strip():
                        raise ValueError("All_freqs_SUBTLEX is blank")
                    zipf_value = self._finite_number(row[14], "Zipf-value")
                except ValueError as error:
                    malformed_rows += 1
                    problems.append(f"row {source_row}: {error}")
                    continue

                range_errors = []
                if source_term != source_term.strip():
                    range_errors.append("Word contains leading or trailing whitespace")
                if " " in source_term:
                    range_errors.append("Word contains whitespace")
                if frequency_count < 1:
                    range_errors.append("FREQcount must be at least 1")
                if not 1 <= contextual_diversity_count <= 8_388:
                    range_errors.append("CDcount is outside 1-8,388")
                if not 0 <= lowercase_frequency_count <= frequency_count:
                    range_errors.append("FREQlow is outside 0-FREQcount")
                if not (
                    0
                    <= lowercase_contextual_diversity_count
                    <= contextual_diversity_count
                ):
                    range_errors.append("Cdlow is outside 0-CDcount")
                if frequency_per_million <= 0:
                    range_errors.append("SUBTLWF must be positive")
                if log10_frequency <= 0:
                    range_errors.append("Lg10WF must be positive")
                if not 0 < contextual_diversity_percent <= 100:
                    range_errors.append("SUBTLCD is outside 0-100")
                if log10_contextual_diversity <= 0:
                    range_errors.append("Lg10CD must be positive")
                if (
                    dominant_source_pos_frequency is not None
                    and dominant_source_pos_frequency < 1
                ):
                    range_errors.append("Freq_dom_PoS_SUBTLEX must be positive")
                if (
                    dominant_source_pos_proportion is not None
                    and not 0 < dominant_source_pos_proportion <= 1
                ):
                    range_errors.append("Percentage_dom_PoS is outside 0-1")
                if not 1 <= zipf_value <= 8:
                    range_errors.append("Zipf-value is outside the documented scale")
                if not math.isclose(
                    log10_frequency,
                    math.log10(frequency_count + 1),
                    rel_tol=1e-12,
                    abs_tol=1e-12,
                ):
                    range_errors.append("Lg10WF disagrees with FREQcount")
                if not math.isclose(
                    log10_contextual_diversity,
                    math.log10(contextual_diversity_count + 1),
                    rel_tol=1e-12,
                    abs_tol=1e-12,
                ):
                    range_errors.append("Lg10CD disagrees with CDcount")
                is_duplicate = lookup_form in entries
                if is_duplicate:
                    duplicate_keys += 1
                    problems.append(
                        f"row {source_row}: duplicate normalized key {lookup_form!r}"
                    )
                if range_errors:
                    out_of_range_values += 1
                    problems.append(
                        f"row {source_row}: " + "; ".join(range_errors)
                    )
                if is_duplicate or range_errors:
                    continue

                entries[lookup_form] = SubtlexUsEntry(
                    source_term=source_term,
                    lookup_form=lookup_form,
                    source_row=source_row,
                    frequency_count=frequency_count,
                    contextual_diversity_count=contextual_diversity_count,
                    lowercase_frequency_count=lowercase_frequency_count,
                    lowercase_contextual_diversity_count=(
                        lowercase_contextual_diversity_count
                    ),
                    frequency_per_million=frequency_per_million,
                    log10_frequency=log10_frequency,
                    contextual_diversity_percent=contextual_diversity_percent,
                    log10_contextual_diversity=log10_contextual_diversity,
                    dominant_source_pos=dominant_source_pos,
                    dominant_source_pos_frequency=dominant_source_pos_frequency,
                    dominant_source_pos_proportion=(
                        dominant_source_pos_proportion
                    ),
                    all_source_pos=all_source_pos,
                    all_source_pos_frequencies=all_source_pos_frequencies,
                    zipf_value=zipf_value,
                )

            errors: list[str] = []
            if blank_terms or malformed_rows or duplicate_keys or out_of_range_values:
                errors.append(
                    "The workbook contains structural problems; no partial "
                    "SUBTLEX-US resource was activated."
                )
            validation = SubtlexUsValidation(
                source_path=path.resolve(),
                source_sha256=source_sha256,
                total_rows=total_rows,
                usable_entries=len(entries),
                blank_terms=blank_terms,
                malformed_rows=malformed_rows,
                duplicate_keys=duplicate_keys,
                out_of_range_values=out_of_range_values,
                errors=tuple(errors),
            )
            if not validation.is_valid:
                detail = " | ".join(problems[:12])
                if len(problems) > 12:
                    detail += f" | {len(problems) - 12} additional problem(s)"
                raise SubtlexUsAdapterError(validation.errors[0], detail)
            return SubtlexUsLexicon.create(entries, validation)
        finally:
            workbook.close()
