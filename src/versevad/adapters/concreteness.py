"""Read-only adapter for the Brysbaert et al. concreteness workbook."""

from __future__ import annotations

import hashlib
import math
from dataclasses import dataclass
from pathlib import Path
from types import MappingProxyType
from typing import Mapping

from openpyxl import load_workbook

from versevad.normalization import normalize_lookup


REQUIRED_COLUMNS = (
    "Word",
    "Bigram",
    "Conc.M",
    "Conc.SD",
    "Unknown",
    "Total",
    "Percent_known",
    "SUBTLEX",
)


class ConcretenessAdapterError(RuntimeError):
    """Plain-language adapter failure that confirms the source stayed unchanged."""

    def __init__(self, message: str, technical_detail: str = "") -> None:
        super().__init__(message)
        self.technical_detail = technical_detail
        self.data_changed = False


@dataclass(frozen=True)
class ConcretenessEntry:
    source_term: str
    lookup_form: str
    source_row: int
    is_multiword: bool
    mean: float
    standard_deviation: float
    unknown_count: int
    rater_count: int
    percent_known: float
    subtlex_count: int

    @property
    def word_count(self) -> int:
        return len(self.lookup_form.split())


@dataclass(frozen=True)
class ConcretenessValidation:
    source_path: Path
    source_sha256: str
    total_rows: int
    usable_entries: int
    phrase_entries: int
    blank_terms: int
    malformed_rows: int
    duplicate_keys: int
    out_of_range_values: int
    errors: tuple[str, ...] = ()
    warnings: tuple[str, ...] = ()

    @property
    def is_valid(self) -> bool:
        return not self.errors


@dataclass(frozen=True)
class ConcretenessLexicon:
    entries: Mapping[str, ConcretenessEntry]
    validation: ConcretenessValidation

    @classmethod
    def create(
        cls,
        entries: Mapping[str, ConcretenessEntry],
        validation: ConcretenessValidation,
    ) -> ConcretenessLexicon:
        return cls(MappingProxyType(dict(entries)), validation)

    def lookup(self, lookup_form: str) -> ConcretenessEntry | None:
        return self.entries.get(lookup_form)

    @property
    def phrase_lengths(self) -> tuple[int, ...]:
        return tuple(
            sorted(
                {
                    entry.word_count
                    for entry in self.entries.values()
                    if entry.is_multiword
                },
                reverse=True,
            )
        )


class BrysbaertConcretenessAdapter:
    """Parse the exact supplementary workbook without saving or rewriting it."""

    adapter_version = "1.0.0"

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _finite_number(value: object, label: str) -> float:
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise ValueError(f"{label} is not numeric")
        number = float(value)
        if not math.isfinite(number):
            raise ValueError(f"{label} is not finite")
        return number

    @classmethod
    def _integer(cls, value: object, label: str) -> int:
        number = cls._finite_number(value, label)
        if not number.is_integer():
            raise ValueError(f"{label} is not an integer")
        return int(number)

    def load(self, source_path: Path | str) -> ConcretenessLexicon:
        path = Path(source_path)
        if not path.is_file():
            raise ConcretenessAdapterError(
                "The concreteness ratings workbook was not found.",
                f"Expected a readable local file at {path}.",
            )
        try:
            source_sha256 = self._sha256(path)
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=False,
                keep_links=False,
            )
        except Exception as error:
            raise ConcretenessAdapterError(
                "The concreteness ratings workbook could not be opened.",
                f"{type(error).__name__}: {error}",
            ) from error

        try:
            if "Sheet1" not in workbook.sheetnames:
                raise ConcretenessAdapterError(
                    "The concreteness workbook does not contain the expected data sheet.",
                    "Expected a worksheet named Sheet1.",
                )
            sheet = workbook["Sheet1"]
            rows = sheet.iter_rows(values_only=True)
            try:
                header = tuple(next(rows))
            except StopIteration as error:
                raise ConcretenessAdapterError(
                    "The concreteness workbook is empty.",
                    "Sheet1 contains no header row.",
                ) from error
            if header != REQUIRED_COLUMNS:
                missing = [
                    column for column in REQUIRED_COLUMNS if column not in header
                ]
                raise ConcretenessAdapterError(
                    "The concreteness workbook does not have the expected columns.",
                    (
                        f"Expected exactly {REQUIRED_COLUMNS!r}; found {header!r}. "
                        f"Missing: {missing!r}."
                    ),
                )

            entries: dict[str, ConcretenessEntry] = {}
            total_rows = 0
            phrase_entries = 0
            blank_terms = 0
            malformed_rows = 0
            duplicate_keys = 0
            out_of_range_values = 0
            problems: list[str] = []

            for source_row, row in enumerate(rows, start=2):
                if not row or all(value is None for value in row):
                    continue
                total_rows += 1
                if len(row) != len(REQUIRED_COLUMNS):
                    malformed_rows += 1
                    problems.append(
                        f"row {source_row}: expected 8 cells, found {len(row)}"
                    )
                    continue
                term_value = row[0]
                if not isinstance(term_value, str) or not term_value.strip():
                    blank_terms += 1
                    problems.append(f"row {source_row}: blank or non-text Word value")
                    continue
                source_term = term_value
                lookup_form = normalize_lookup(source_term)
                try:
                    bigram = self._integer(row[1], "Bigram")
                    mean = self._finite_number(row[2], "Conc.M")
                    standard_deviation = self._finite_number(row[3], "Conc.SD")
                    unknown_count = self._integer(row[4], "Unknown")
                    rater_count = self._integer(row[5], "Total")
                    percent_known = self._finite_number(row[6], "Percent_known")
                    subtlex_count = self._integer(row[7], "SUBTLEX")
                except ValueError as error:
                    malformed_rows += 1
                    problems.append(f"row {source_row}: {error}")
                    continue

                range_errors = []
                if bigram not in {0, 1}:
                    range_errors.append("Bigram must be 0 or 1")
                if not 1 <= mean <= 5:
                    range_errors.append("Conc.M is outside the source 1-5 scale")
                if standard_deviation < 0:
                    range_errors.append("Conc.SD is negative")
                if rater_count <= 0 or not 0 <= unknown_count <= rater_count:
                    range_errors.append("Unknown/Total counts are inconsistent")
                if not 0 <= percent_known <= 1:
                    range_errors.append("Percent_known is outside 0-1")
                if subtlex_count < 0:
                    range_errors.append("SUBTLEX is negative")
                expected_known = (
                    1 - (unknown_count / rater_count) if rater_count else None
                )
                if expected_known is not None and not math.isclose(
                    percent_known,
                    expected_known,
                    rel_tol=1e-9,
                    abs_tol=1e-9,
                ):
                    range_errors.append(
                        "Percent_known does not agree with Unknown and Total"
                    )
                has_space = " " in source_term
                if bool(bigram) != has_space:
                    range_errors.append(
                        "Bigram flag does not agree with the spaced source term"
                    )
                if source_term != source_term.strip():
                    range_errors.append(
                        "Word contains leading or trailing whitespace"
                    )
                is_duplicate = lookup_form in entries
                if is_duplicate:
                    duplicate_keys += 1
                    problems.append(
                        f"row {source_row}: duplicate normalized key {lookup_form!r}"
                    )
                if range_errors:
                    out_of_range_values += 1
                    problems.append(
                        f"row {source_row}: " + "; ".join(range_errors)
                    )
                if is_duplicate or range_errors:
                    continue

                entry = ConcretenessEntry(
                    source_term=source_term,
                    lookup_form=lookup_form,
                    source_row=source_row,
                    is_multiword=bool(bigram),
                    mean=mean,
                    standard_deviation=standard_deviation,
                    unknown_count=unknown_count,
                    rater_count=rater_count,
                    percent_known=percent_known,
                    subtlex_count=subtlex_count,
                )
                entries[lookup_form] = entry
                phrase_entries += int(entry.is_multiword)

            errors: list[str] = []
            if blank_terms or malformed_rows or duplicate_keys or out_of_range_values:
                errors.append(
                    "The workbook contains structural problems; no partial "
                    "concreteness resource was activated."
                )
            validation = ConcretenessValidation(
                source_path=path.resolve(),
                source_sha256=source_sha256,
                total_rows=total_rows,
                usable_entries=len(entries),
                phrase_entries=phrase_entries,
                blank_terms=blank_terms,
                malformed_rows=malformed_rows,
                duplicate_keys=duplicate_keys,
                out_of_range_values=out_of_range_values,
                errors=tuple(errors),
            )
            if not validation.is_valid:
                detail = " | ".join(problems[:12])
                if len(problems) > 12:
                    detail += f" | {len(problems) - 12} additional problem(s)"
                raise ConcretenessAdapterError(
                    validation.errors[0],
                    detail,
                )
            return ConcretenessLexicon.create(entries, validation)
        finally:
            workbook.close()
