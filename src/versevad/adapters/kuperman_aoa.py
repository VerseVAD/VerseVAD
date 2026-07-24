"""Read-only adapter for the official Kuperman et al. AoA supplement."""

from __future__ import annotations

import hashlib
import math
from dataclasses import dataclass
from pathlib import Path
from types import MappingProxyType
from typing import Mapping

from openpyxl import load_workbook

from versevad.normalization import normalize_lookup


REQUIRED_COLUMNS = (
    "Word",
    "OccurTotal",
    "OccurNum",
    "Freq_pm",
    "Rating.Mean",
    "Rating.SD",
    "Dunno",
)


class KupermanAoAAdapterError(RuntimeError):
    """Plain-language adapter failure that confirms the source stayed unchanged."""

    def __init__(self, message: str, technical_detail: str = "") -> None:
        super().__init__(message)
        self.technical_detail = technical_detail
        self.data_changed = False


@dataclass(frozen=True)
class KupermanAoAEntry:
    """One source row, including unavailable ratings and response evidence."""

    source_term: str
    lookup_form: str
    source_row: int
    occurrence_total: int
    numeric_response_count: int
    frequency_per_million: float | None
    mean_age: float | None
    standard_deviation: float | None
    source_dunno_value: float

    @property
    def unknown_response_count(self) -> int:
        return self.occurrence_total - self.numeric_response_count

    @property
    def numeric_response_proportion(self) -> float:
        return self.numeric_response_count / self.occurrence_total


@dataclass(frozen=True)
class KupermanAoAValidation:
    source_path: Path
    source_sha256: str
    total_rows: int
    rated_entries: int
    unavailable_entries: int
    blank_terms: int
    malformed_rows: int
    duplicate_keys: int
    out_of_range_values: int
    errors: tuple[str, ...] = ()
    warnings: tuple[str, ...] = ()

    @property
    def is_valid(self) -> bool:
        return not self.errors


@dataclass(frozen=True)
class KupermanAoALexicon:
    entries: Mapping[str, KupermanAoAEntry]
    validation: KupermanAoAValidation

    @classmethod
    def create(
        cls,
        entries: Mapping[str, KupermanAoAEntry],
        validation: KupermanAoAValidation,
    ) -> KupermanAoALexicon:
        return cls(MappingProxyType(dict(entries)), validation)

    def lookup(self, lookup_form: str) -> KupermanAoAEntry | None:
        return self.entries.get(lookup_form)


class KupermanAoAAdapter:
    """Parse the exact Springer erratum workbook without saving or rewriting it."""

    adapter_version = "1.0.0"

    @staticmethod
    def _sha256(path: Path) -> str:
        digest = hashlib.sha256()
        with path.open("rb") as source:
            for chunk in iter(lambda: source.read(1024 * 1024), b""):
                digest.update(chunk)
        return digest.hexdigest()

    @staticmethod
    def _finite_number(value: object, label: str) -> float:
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            raise ValueError(f"{label} is not numeric")
        number = float(value)
        if not math.isfinite(number):
            raise ValueError(f"{label} is not finite")
        return number

    @classmethod
    def _integer(cls, value: object, label: str) -> int:
        number = cls._finite_number(value, label)
        if not number.is_integer():
            raise ValueError(f"{label} is not an integer")
        return int(number)

    @classmethod
    def _optional_frequency(cls, value: object) -> float | None:
        if value == "#N/A":
            return None
        return cls._finite_number(value, "Freq_pm")

    @classmethod
    def _optional_rating(cls, value: object, label: str) -> float | None:
        if value == "NA":
            return None
        return cls._finite_number(value, label)

    def load(self, source_path: Path | str) -> KupermanAoALexicon:
        path = Path(source_path)
        if not path.is_file():
            raise KupermanAoAAdapterError(
                "The Kuperman age-of-acquisition workbook was not found.",
                f"Expected a readable local file at {path}.",
            )
        try:
            source_sha256 = self._sha256(path)
            workbook = load_workbook(
                path,
                read_only=True,
                data_only=True,
                keep_links=False,
            )
        except Exception as error:
            raise KupermanAoAAdapterError(
                "The Kuperman age-of-acquisition workbook could not be opened.",
                f"{type(error).__name__}: {error}",
            ) from error

        try:
            if "Sheet1" not in workbook.sheetnames:
                raise KupermanAoAAdapterError(
                    "The Kuperman workbook does not contain the expected data sheet.",
                    "Expected a worksheet named Sheet1.",
                )
            sheet = workbook["Sheet1"]
            rows = sheet.iter_rows(values_only=True)
            try:
                header = tuple(next(rows))
            except StopIteration as error:
                raise KupermanAoAAdapterError(
                    "The Kuperman workbook is empty.",
                    "Sheet1 contains no header row.",
                ) from error
            if header != REQUIRED_COLUMNS:
                missing = [
                    column for column in REQUIRED_COLUMNS if column not in header
                ]
                raise KupermanAoAAdapterError(
                    "The Kuperman workbook does not have the expected columns.",
                    (
                        f"Expected exactly {REQUIRED_COLUMNS!r}; found {header!r}. "
                        f"Missing: {missing!r}."
                    ),
                )

            entries: dict[str, KupermanAoAEntry] = {}
            total_rows = 0
            rated_entries = 0
            unavailable_entries = 0
            blank_terms = 0
            malformed_rows = 0
            duplicate_keys = 0
            out_of_range_values = 0
            problems: list[str] = []

            for source_row, row in enumerate(rows, start=2):
                if not row or all(value is None for value in row):
                    continue
                total_rows += 1
                if len(row) != len(REQUIRED_COLUMNS):
                    malformed_rows += 1
                    problems.append(
                        f"row {source_row}: expected 7 cells, found {len(row)}"
                    )
                    continue
                term_value = row[0]
                if not isinstance(term_value, str) or not term_value.strip():
                    blank_terms += 1
                    problems.append(f"row {source_row}: blank or non-text Word value")
                    continue
                source_term = term_value
                lookup_form = normalize_lookup(source_term)
                try:
                    occurrence_total = self._integer(row[1], "OccurTotal")
                    numeric_response_count = self._integer(row[2], "OccurNum")
                    frequency_per_million = self._optional_frequency(row[3])
                    mean_age = self._optional_rating(row[4], "Rating.Mean")
                    standard_deviation = self._optional_rating(
                        row[5], "Rating.SD"
                    )
                    source_dunno_value = self._finite_number(row[6], "Dunno")
                except ValueError as error:
                    malformed_rows += 1
                    problems.append(f"row {source_row}: {error}")
                    continue

                range_errors = []
                if source_term != source_term.strip():
                    range_errors.append("Word contains leading or trailing whitespace")
                if any(character.isspace() for character in source_term):
                    range_errors.append("Word contains whitespace")
                if not lookup_form:
                    range_errors.append("Word has an empty normalized lookup key")
                if occurrence_total < 1:
                    range_errors.append("OccurTotal must be at least 1")
                if not 0 <= numeric_response_count <= occurrence_total:
                    range_errors.append("OccurNum is outside 0-OccurTotal")
                if (
                    frequency_per_million is not None
                    and frequency_per_million <= 0
                ):
                    range_errors.append("Freq_pm must be positive when available")
                if mean_age is not None and not 0 <= mean_age <= 25:
                    range_errors.append("Rating.Mean is outside the retained 0-25 range")
                if (
                    standard_deviation is not None
                    and not 0 <= standard_deviation <= 25
                ):
                    range_errors.append("Rating.SD is outside 0-25")
                if mean_age is None and numeric_response_count != 0:
                    range_errors.append(
                        "Rating.Mean is unavailable although OccurNum is nonzero"
                    )
                if mean_age is not None and numeric_response_count == 0:
                    range_errors.append(
                        "Rating.Mean is numeric although OccurNum is zero"
                    )
                if standard_deviation is None and numeric_response_count >= 2:
                    range_errors.append(
                        "Rating.SD is unavailable despite at least two numeric responses"
                    )
                if standard_deviation is not None and numeric_response_count < 2:
                    range_errors.append(
                        "Rating.SD is numeric with fewer than two numeric responses"
                    )
                if not 0 <= source_dunno_value <= 1:
                    range_errors.append("Dunno is outside 0-1")
                if occurrence_total >= 1 and not math.isclose(
                    source_dunno_value,
                    numeric_response_count / occurrence_total,
                    rel_tol=1e-12,
                    abs_tol=1e-12,
                ):
                    range_errors.append(
                        "Dunno disagrees with OccurNum / OccurTotal"
                    )
                is_duplicate = lookup_form in entries
                if is_duplicate:
                    duplicate_keys += 1
                    problems.append(
                        f"row {source_row}: duplicate normalized key {lookup_form!r}"
                    )
                if range_errors:
                    out_of_range_values += 1
                    problems.append(
                        f"row {source_row}: " + "; ".join(range_errors)
                    )
                if is_duplicate or range_errors:
                    continue

                entry = KupermanAoAEntry(
                    source_term=source_term,
                    lookup_form=lookup_form,
                    source_row=source_row,
                    occurrence_total=occurrence_total,
                    numeric_response_count=numeric_response_count,
                    frequency_per_million=frequency_per_million,
                    mean_age=mean_age,
                    standard_deviation=standard_deviation,
                    source_dunno_value=source_dunno_value,
                )
                entries[lookup_form] = entry
                if mean_age is None:
                    unavailable_entries += 1
                else:
                    rated_entries += 1

            errors: list[str] = []
            if blank_terms or malformed_rows or duplicate_keys or out_of_range_values:
                errors.append(
                    "The workbook contains structural problems; no partial "
                    "Kuperman AoA resource was activated."
                )
            warning_tuple = (
                (
                    f"{unavailable_entries} source entries have no numeric "
                    "Rating.Mean and remain unavailable.",
                )
                if unavailable_entries
                else ()
            )
            validation = KupermanAoAValidation(
                source_path=path.resolve(),
                source_sha256=source_sha256,
                total_rows=total_rows,
                rated_entries=rated_entries,
                unavailable_entries=unavailable_entries,
                blank_terms=blank_terms,
                malformed_rows=malformed_rows,
                duplicate_keys=duplicate_keys,
                out_of_range_values=out_of_range_values,
                errors=tuple(errors),
                warnings=warning_tuple,
            )
            if not validation.is_valid:
                detail = " | ".join(problems[:12])
                if len(problems) > 12:
                    detail += f" | {len(problems) - 12} additional problem(s)"
                raise KupermanAoAAdapterError(validation.errors[0], detail)
            return KupermanAoALexicon.create(entries, validation)
        finally:
            workbook.close()
