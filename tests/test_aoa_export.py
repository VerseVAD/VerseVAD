from __future__ import annotations

import csv
import hashlib
import io
from pathlib import Path

from openpyxl import Workbook

from versevad.core import ModuleInput, ResourceSpec
from versevad.exports.aoa import (
    export_aoa_bundle,
    export_aoa_distribution_csv,
    export_aoa_summary_csv,
    export_aoa_terms_csv,
    export_aoa_token_audit_csv,
)
from versevad.lexical_semantic.aoa import AoAConfiguration, AoAModule
from versevad.adapters.kuperman_aoa import REQUIRED_COLUMNS
from versevad.preprocessing import create_text_document


def _result(tmp_path: Path, preprocessor):
    source = tmp_path / "aoa.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(REQUIRED_COLUMNS)
    sheet.append(("stone", 20, 20, 10.0, 4.0, 1.0, 1.0))
    sheet.append(("mystery", 20, 10, 1.0, 12.0, 2.0, 0.5))
    workbook.create_sheet("Sheet2")
    workbook.create_sheet("Sheet3")
    workbook.save(source)
    spec = ResourceSpec(
        resource_id="synthetic-aoa-export",
        display_name="Synthetic AoA export fixture",
        relative_path=source.name,
        version="synthetic-v1",
        accepted_sha256=(hashlib.sha256(source.read_bytes()).hexdigest(),),
        citation="Constructed fixture.",
        license_notice="Synthetic data.",
    )
    poem = preprocessor.process_document(
        create_text_document(
            "aoa-export",
            "AoA export",
            "stone mystery quorvax",
        )
    )
    return AoAModule(tmp_path, resource_spec=spec).analyze_detailed(
        ModuleInput.from_poem_document(poem),
        AoAConfiguration(exclude_proper_nouns=False),
    )


def _csv_rows(content: bytes) -> list[dict[str, str]]:
    return list(
        csv.DictReader(
            io.StringIO(content.decode("utf-8-sig"), newline="")
        )
    )


def test_aoa_csv_exports_preserve_missing_values_and_response_evidence(
    tmp_path: Path,
    preprocessor,
) -> None:
    result = _result(tmp_path, preprocessor)

    summary = _csv_rows(export_aoa_summary_csv(result))
    audit = _csv_rows(export_aoa_token_audit_csv(result))
    terms = _csv_rows(export_aoa_terms_csv(result))
    bands = _csv_rows(export_aoa_distribution_csv(result))

    mean_row = next(row for row in summary if row["metric"] == "mean_normative_aoa")
    assert float(mean_row["value"]) == 8.0
    assert "source mean age" in mean_row["unit_or_scale"]
    unmatched = next(row for row in audit if row["surface_form"] == "quorvax")
    assert unmatched["included"] == "False"
    assert unmatched["mean_age"] == ""
    mystery = next(row for row in terms if row["source_term"] == "mystery")
    assert mystery["source_numeric_response_count"] == "10"
    assert mystery["source_numeric_response_proportion"] == "0.5"
    assert sum(int(row["token_count"]) for row in bands) == 2


def test_aoa_word_report_and_bundle_are_complete_and_deterministic(
    tmp_path: Path,
    preprocessor,
) -> None:
    result = _result(tmp_path, preprocessor)

    bundle = export_aoa_bundle(result)
    second = export_aoa_bundle(result)

    assert bundle["aoa_report.docx"] == second["aoa_report.docx"]
    assert result.module_result.module_name == "age_of_acquisition"
    assert result.summary.statistics.mean == 8.0
    assert any(
        warning.code == "aoa_non_diagnostic"
        for warning in result.module_result.warnings
    )
    assert set(bundle) == {
        "aoa_summary.csv",
        "aoa_distribution.csv",
        "aoa_by_structure.csv",
        "aoa_by_pos.csv",
        "aoa_terms.csv",
        "aoa_relationships.csv",
        "aoa_token_audit.csv",
        "aoa_manifest.csv",
        "aoa_report.docx",
    }
    assert bundle["aoa_report.docx"].startswith(b"PK")
    assert not any(name.endswith(".json") for name in bundle)
