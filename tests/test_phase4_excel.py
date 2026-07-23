from io import BytesIO

from openpyxl import load_workbook

from versevad.db import (
    CorpusMetricRecord,
    CorpusTextRecord,
    ProjectRecord,
    UnmatchedQcRecord,
)
from versevad.exports.corpus_excel import (
    CORPUS_WORKBOOK_API_VERSION,
    build_corpus_workbook,
)


def test_corpus_excel_is_readable_and_contains_both_collection_views() -> None:
    project = ProjectRecord("project-1", "Jeffers volume", "", "Scholar", "now", "now")
    text = CorpusTextRecord(
        "text-1",
        "version-1",
        project.project_id,
        "Poem",
        "poem.txt",
        "poem.txt",
        "Robinson Jeffers",
        "Volume",
        "1925",
        "lyric",
        "",
        {},
        "Bright.",
        "abc123",
        "now",
        "now",
    )

    def metric(name: str, value: float, *, weighting: str = "token") -> CorpusMetricRecord:
        return CorpusMetricRecord(
            "run-1",
            text.text_id,
            text.text_version_id,
            text.title,
            text.author,
            text.collection,
            text.date_label,
            text.genre,
            "vad-test",
            "VAD test",
            "vad",
            name,
            "valence",
            "",
            weighting,
            "normalized_0_1" if name == "vad_mean" else "midpoint_deviation_sum",
            "1 included matched observation",
            value,
            1,
            1,
            1,
            1.0,
            "now",
        )

    metrics = (
        metric("vad_mean", 0.875),
        metric("vad_mean", 0.875, weighting="type"),
        metric("vad_absolute_midpoint_load", 0.375),
    )
    unmatched = (
        UnmatchedQcRecord(
            project.project_id,
            text.text_id,
            text.title,
            "vad-test",
            "VAD test",
            "mystery",
            "mystery",
            1,
            "NOUN",
            "mystery",
            1,
            "Bright mystery.",
            "needs mapping",
            "Review historical sense.",
            "",
            "note-1",
            "now",
        ),
    )
    content = build_corpus_workbook(project, (text,), metrics, unmatched)
    workbook = load_workbook(BytesIO(content), data_only=False)
    assert workbook.sheetnames == [
        "START HERE",
        "Corpus Profiles",
        "Work VAD",
        "Cumulative Load",
        "Coverage and Emotion",
        "Unmatched QC",
        "Text Metadata",
    ]
    assert workbook["START HERE"]["A1"].value == "VerseVAD corpus workbook"
    profile_headers = [cell.value for cell in workbook["Corpus Profiles"][4]]
    assert "Token-weighted volume mean" in profile_headers
    assert "Work-weighted volume mean" in profile_headers
    assert workbook["Corpus Profiles"]["I5"].value == 0.875
    assert workbook["Corpus Profiles"]["B5"].value == "All Matched"
    assert workbook["Corpus Profiles"]["I5"].value == 0.875
    assert workbook["Unmatched QC"]["K5"].value == "Review historical sense."
    assert workbook["Text Metadata"]["I5"].value == "abc123"
    coverage_headers = [cell.value for cell in workbook["Coverage and Emotion"][4]]
    assert "Construct" in coverage_headers

    methodology_content = build_corpus_workbook(
        project,
        (text,),
        metrics,
        unmatched,
        methodology={
            "software_version": "0.6.0.dev0",
            "scenario_id": "scenario-test",
            "phrase_policy": "phrase_preferred",
            "minimum_match_requirement": 1,
            "stopword_policy": {
                "mode": "standard",
                "source": "spaCy English STOP_WORDS",
                "library_version": "3.8.14",
                "list_version": "test-list",
                "active_list_sha256": "abc123",
                "protected_words": ("not", "never"),
                "custom_additions": (),
                "custom_removals": (),
            },
        },
    )
    methodology_workbook = load_workbook(
        BytesIO(methodology_content),
        data_only=False,
    )
    assert CORPUS_WORKBOOK_API_VERSION == 4
    assert methodology_workbook.sheetnames[-1] == "Methodology"
    methodology_rows = {
        row[0].value: row[1].value
        for row in methodology_workbook["Methodology"].iter_rows(min_row=5)
        if row[0].value
    }
    assert methodology_rows["Stopword source"] == "spaCy English STOP_WORDS"
    assert methodology_rows["Protected words"] == "not, never"

    reviewed_content = build_corpus_workbook(
        project,
        (text,),
        metrics,
        unmatched,
        methodology={
            "scenario_id": "scenario-test",
            "scenario_version_id": "scenario-version-2",
            "review_decisions": ({"decision_revision_id": "revision-1"},),
        },
        review_decisions=(
            {
                "decision_id": "decision-1",
                "decision_revision_id": "revision-1",
                "action": "map",
                "scope": "work",
                "lexicon_id": "vad-test",
                "source_form": "o'er",
                "mapping_target": "over",
                "project_id": project.project_id,
                "text_id": text.text_id,
                "text_version_id": text.text_version_id,
                "token_position": None,
                "risk_category": "unmatched",
                "rationale": "Edition-level contraction mapping.",
            },
        ),
    )
    reviewed_workbook = load_workbook(BytesIO(reviewed_content), data_only=False)
    assert "Review Decisions" in reviewed_workbook.sheetnames
    assert reviewed_workbook["Review Decisions"]["F5"].value == "o'er"
    assert reviewed_workbook["Review Decisions"]["G5"].value == "over"

    pos_content = build_corpus_workbook(
        project,
        (text,),
        metrics,
        unmatched,
        part_of_speech_rows=(
            {
                "Scope": "Work",
                "Work": "Poem",
                "Collection": "Volume",
                "Universal POS tag": "ADJ",
                "Part of speech": "Adjective",
                "Token count": 1,
                "Share of lexical tokens": 1.0,
                "Unique normalized types": 1,
                "Examples": "bright",
                "Lexical-token denominator": 1,
                "Model": "en_core_web_sm test",
            },
        ),
    )
    pos_workbook = load_workbook(BytesIO(pos_content), data_only=False)
    assert "Part of Speech" in pos_workbook.sheetnames
    assert pos_workbook["Part of Speech"]["E5"].value == "Adjective"
    assert pos_workbook["Part of Speech"]["G5"].value == 1.0
